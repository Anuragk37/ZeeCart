from django.contrib import messages
from django.shortcuts import redirect, render
from django.contrib.auth import authenticate, login,logout
from django.contrib.auth.decorators import login_required,user_passes_test
from django.http import HttpResponse, JsonResponse
from django.core.serializers.json import DjangoJSONEncoder
from openpyxl import Workbook



from .forms import*
from .models import*
from userhome.models import *
from user_order.models import *
from coupon_banner.models import Banner
from django.db.models import Sum
from django.utils import timezone
from datetime import timedelta

import plotly.express as px
import pandas as pd


# Create your views here.
@login_required(login_url='admin_login')
def dashboard(request):
    if request.user.is_admin:
        user_count=NewUser.objects.all().count()
        orders_count=Order.objects.all().count()
        total_revenue = Order.objects.aggregate(total_revenue=Sum('total_amount'))['total_revenue']

        order_delevered=OrderItems.objects.filter(order_status='Delivered').count()
        order_pending=OrderItems.objects.exclude(order_status__in = ('Devivered','Canceled')).count()

        orders = Order.objects.all()

        order_items=OrderItems.objects.all().count()

        # ----------------- graph plot---------------->
        data = {
            'order_date': [order.order_date.date() for order in orders],
            'total_amount': [order.total_amount for order in orders],
        }

        df = pd.DataFrame(data)

        df['order_date'] = pd.to_datetime(df['order_date'])
        df.set_index('order_date', inplace=True)

        df_daily = df.resample('D').sum()

        df_daily.reset_index(inplace=True)


        fig = px.line(df_daily, x='order_date', y='total_amount', labels={'total_amount': 'Total Amount'}, title='Daily Order Amount')

        fig.update_traces(
            line=dict(color='#7076F0', width=2),  
            marker=dict(size=8, color='blue', line=dict(color='white', width=2)),  
        )

        # Customize the layout
        fig.update_layout(
            xaxis=dict(title='Order Date'),  # X-axis title
            yaxis=dict(title='Total Amount'),  # Y-axis title
            template='seaborn',  # Change the chart template (options: "plotly", "plotly_white", "plotly_dark", etc.)
        )
        # Convert the plot to HTML
        plot_html = fig.to_html(full_html=False)

        # -------------- graph plot end ------------------->
        
        context={
            'user_count':user_count,
            'orders_count':orders_count,
            'total_revenue':total_revenue,
            'plot_html':plot_html,
            'order_delevered':order_delevered,
            'order_pending':order_pending,
            'order_items':order_items
        }
        return render(request,'adminhome/dashboard.html',context)
    else:
        return redirect('admin_login')

def admin_login(request):
    if request.method=='POST':
        email=request.POST['email']
        password=request.POST['password']
        user=authenticate(request,email=email,password=password)
        if user is not None and user.is_admin:
            login(request,user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid credentials. Please try again.')
            return render(request,'adminhome/signin.html')
    return render(request,'adminhome/signin.html')


def admin_logout(request):
    logout(request)
    return redirect('admin_login')

@login_required(login_url='admin_login')
@user_passes_test(lambda u: u.is_admin)
def add_banner(request):
    if request.user.is_authenticated:
        form=BannerForm()
        if request.method=='POST':
            form=BannerForm(request.POST,request.FILES)
            if form.is_valid():
                form.save()
                return redirect('add_banner')
            else:
                return redirect('add_banner')
        return render(request,'adminhome/add-banner.html',{'form':form})
    else:
        messages.error(request,'Your are not an admin')
        return redirect('home')

@login_required(login_url='admin_login')
def sales_report(request):
    if request.user.is_admin:
        today = timezone.now().date()
        todays_order=Order.objects.filter(order_date__date = today)
        orders=todays_order.count()
        total_revenue = Order.objects.filter(order_date__date = today).aggregate(total_revenue=Sum('total_amount'))['total_revenue']
        today_order_items=OrderItems.objects.filter(order__in=todays_order).count()
        

        top_products = OrderItems.objects.filter(order__in=todays_order).values(
            'product_varient__product__product_name'
        ).annotate(
            total_quantity=Sum('quantity')
        ).order_by('-total_quantity')[:5]

        today_order_products=OrderItems.objects.filter(order__in=todays_order)

        week_ago = today - timedelta(days=7)
        week_order = Order.objects.filter(order_date__range=[week_ago, today + timedelta(days=1)])
        week_order_count = week_order.count()
        week_revenue = week_order.aggregate(week_revenue=Sum('total_amount'))['week_revenue']
        weekly_order_items = OrderItems.objects.filter(order__in=week_order).count()

        week_top_products = OrderItems.objects.filter(order__in=week_order).values(
            'product_varient__product__product_name'
        ).annotate(
            total_quantity=Sum('quantity')
        ).order_by('-total_quantity')[:5]

        weekly_order_products=OrderItems.objects.filter(order__in=week_order)


        month_start = today.replace(day=1)
        month_order=Order.objects.filter(order_date__range=[month_start, today + timedelta(days=1)])
        month_order_count=month_order.count()
        month_revenue = Order.objects.filter(order_date__range=[month_start, today]).aggregate(month_revenue=Sum('total_amount'))['month_revenue']
        monthly_order_items=OrderItems.objects.filter(order__in=month_order).count()

        month_top_products = OrderItems.objects.filter(order__in=month_order).values(
            'product_varient__product__product_name','quantity'
        ).annotate(
            total_quantity=Sum('quantity')
        ).order_by('-total_quantity')[:5]

        monthly_order_products=OrderItems.objects.filter(order__in=month_order)

        year_start = today.replace(month=1, day=1)
        year_order = Order.objects.filter(order_date__range=[year_start, today + timedelta(days=1)])
        year_order_count = year_order.count()
        year_revenue = year_order.aggregate(year_revenue=Sum('total_amount'))['year_revenue']
        yearly_order_items = OrderItems.objects.filter(order__in=year_order).count()

        year_top_products = OrderItems.objects.filter(order__in=year_order).values(
            'product_varient__product__product_name'
        ).annotate(
            total_quantity=Sum('quantity')
        ).order_by('-total_quantity')[:5]

        yearly_order_products=OrderItems.objects.filter(order__in=year_order)


    

        context={
            'todays_order':orders,
            'total_revenue':total_revenue,
            'top_products':top_products,
            'today_order_items':today_order_items,
            'today_order_products':today_order_products,

            'week_order_count':week_order_count,
            'week_revenue':week_revenue,
            'weekly_order_items':weekly_order_items,
            'week_top_products':week_top_products,
            'weekly_order_products':weekly_order_products,

            'month_order_count':month_order_count,
            'month_revenue':month_revenue,
            'month_top_products':month_top_products,
            'monthly_order_items':monthly_order_items,
            'monthly_order_products':monthly_order_products,

            'year_order_count':year_order_count,
            'year_revenue':year_revenue,
            'yearly_order_items':yearly_order_items,
            'year_top_products':year_top_products,
            'yearly_order_products':yearly_order_products


        }
        return render(request,'adminhome/sales-report.html',context)
    else:
        return redirect('admin_login')

from django.utils import timezone
from datetime import datetime
from django.http import JsonResponse
from django.core.serializers import serialize

def date_filter(request):
    from_date = timezone.make_aware(datetime.strptime(request.GET.get('from_date'), "%Y-%m-%d"))
    to_date = timezone.make_aware(datetime.strptime(request.GET.get('to_date'), "%Y-%m-%d"))

    orders = Order.objects.filter(order_date__range=[from_date, to_date])
    order_count = orders.count()
    
    revenue = Order.objects.filter(order_date__range=[from_date, to_date]).aggregate(week_revenue=Sum('total_amount'))['week_revenue']

    order_items = OrderItems.objects.filter(order__in=orders).count()
    
    top_products = list(
        OrderItems.objects.filter(order__in=orders).values(
            'product_varient__product__product_name', 'quantity'
        ).annotate(
            total_quantity=Sum('quantity')
        ).order_by('-total_quantity')[:5]
    )

    order_products = OrderItems.objects.filter(order__in=orders)

    order_products_data = order_products.values(
        'order_id',
        'product_varient__product__product_name',
        'product_varient__color__color',  
        'quantity',
        'price',
    )
 
    order_products_list = list(order_products_data)
   
    context = {
        'order_count': order_count,
        'revenue': revenue,
        'order_items': order_items,
        'top_products': top_products,
        'order_products': order_products_list,
    }

    return JsonResponse(context, safe=False)


from django.template.loader import render_to_string
from xhtml2pdf import pisa
@login_required(login_url='admin_login')
def sales_report_pdf(request,time_period):

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if time_period == 'today':
        today = timezone.now().date()
        todays_order=Order.objects.filter(order_date__date = today)
        orders=todays_order.count()
        total_revenue = Order.objects.filter(order_date__date = today).aggregate(total_revenue=Sum('total_amount'))['total_revenue']
        today_order_items=OrderItems.objects.filter(order__in=todays_order).count()
        

        top_products = OrderItems.objects.filter(order__in=todays_order).values(
            'product_varient__product__product_name',
        ).annotate(
            total_quantity=Sum('quantity')
        ).order_by('-total_quantity')[:5]

        today_order_products=OrderItems.objects.filter(order__in=todays_order)

        context={
            'order_count':orders,
            'total_revenue':total_revenue,
            'top_products':top_products,
            'order_items_count':today_order_items,
            'order_products':today_order_products,
        }
    elif time_period =='week':
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        week_order=Order.objects.filter(order_date__range=[week_ago, today])
        week_order_count=week_order.count()
        week_revenue = Order.objects.filter(order_date__range=[week_ago, today]).aggregate(week_revenue=Sum('total_amount'))['week_revenue']
        weekly_order_items=OrderItems.objects.filter(order__in=week_order).count()

        week_top_products = OrderItems.objects.filter(order__in=week_order).values(
            'product_varient__product__product_name',
        ).annotate(
            total_quantity=Sum('quantity')
        ).order_by('-total_quantity')[:5]

        weekly_order_products=OrderItems.objects.filter(order__in=week_order)

        context={
            'order_count':week_order_count,
            'total_revenue':week_revenue,
            'top_products':week_top_products,
            'order_items_count':weekly_order_items,
            'order_products':weekly_order_products,
        }
    
    elif time_period == 'month':
        today = timezone.now().date()
        month_start = today.replace(day=1)
        month_order=Order.objects.filter(order_date__range=[month_start, today])
        month_order_count=month_order.count()
        month_revenue = Order.objects.filter(order_date__range=[month_start, today]).aggregate(month_revenue=Sum('total_amount'))['month_revenue']
        monthly_order_items=OrderItems.objects.filter(order__in=month_order).count()

        month_top_products = OrderItems.objects.filter(order__in=month_order).values(
            'product_varient__product__product_name','quantity'
        ).annotate(
            total_quantity=Sum('quantity')
        ).order_by('-total_quantity')[:5]

        monthly_order_products=OrderItems.objects.filter(order__in=month_order)

        

        context={
            'order_count':month_order_count,
            'total_revenue':month_revenue,
            'top_products':month_top_products,
            'order_items_count':monthly_order_items,
            'order_products':monthly_order_products,

        }
    elif time_period == 'year':
        today = timezone.now().date()
        month_start = today.replace(day=1)
        year_start = today.replace(month=1, day=1)
        year_order = Order.objects.filter(order_date__range=[year_start, today])
        year_order_count = year_order.count()
        year_revenue = year_order.aggregate(year_revenue=Sum('total_amount'))['year_revenue']
        yearly_order_items = OrderItems.objects.filter(order__in=year_order).count()
        year_top_products = OrderItems.objects.filter(order__in=year_order).values(
            'product_varient__product__product_name', 'quantity'
        ).annotate(
            total_quantity=Sum('quantity')
        ).order_by('-total_quantity')[:5]

        yearly_order_products=OrderItems.objects.filter(order__in=year_order)

        context={
            'order_count':year_order_count,
            'total_revenue':year_revenue,
            'top_products':year_top_products,
            'order_items_count':yearly_order_items,
            'order_products':yearly_order_products,
        }
    
    elif from_date and to_date:
        orders = Order.objects.filter(order_date__range=[from_date, to_date])
        order_count=orders.count()
        revenue = Order.objects.filter(order_date__range=[from_date, to_date]).aggregate(week_revenue=Sum('total_amount'))['week_revenue']

        order_items = OrderItems.objects.filter(order__in=orders).count()
        top_products = list(
            OrderItems.objects.filter(order__in=orders).values(
                'product_varient__product__product_name', 'quantity'
            ).annotate(
                total_quantity=Sum('quantity')
            ).order_by('-total_quantity')[:5]
        )
        date_order_products=OrderItems.objects.filter(order__in=orders)
        context={
            'order_count':order_count,
            'total_revenue':revenue,
            'top_products':top_products,
            'order_items_count':order_items,
            'order_products':date_order_products,
        }

    

    template = 'adminhome/sales-report-pdf.html'
    html_string = render_to_string(template, context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename="sales_report.pdf"'

    pisa_status = pisa.CreatePDF(html_string, dest=response )
    
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html_string + '</pre>')
    return response

from openpyxl.styles import Font

def excel_report(request,time_period):
    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    def set_header_styles(cell):
        cell.font = Font(bold=True)

    if time_period == 'today':
        today = timezone.now().date()
        todays_order=Order.objects.filter(order_date__date = today)
        orders=todays_order.count()
        total_revenue = Order.objects.filter(order_date__date = today).aggregate(total_revenue=Sum('total_amount'))['total_revenue']
        today_order_items=OrderItems.objects.filter(order__in=todays_order).count()
        

        top_products = OrderItems.objects.filter(order__in=todays_order).values(
            'product_varient__product__product_name',
        ).annotate(
            total_quantity=Sum('quantity')
        ).order_by('-total_quantity')[:5]

        ws['A1'] = 'Total Order'
        ws['A2'] = 'Total revenue'
        ws['A3'] = 'Total order items'
        ws['B1'] = orders
        ws['B2'] = total_revenue
        ws['B3'] = today_order_items

        for cell in ws['A1:B3']:
            set_header_styles(cell[0])

        ws['A6']='Product'
        ws['B6']='Quantity'

        set_header_styles(ws['A6'])
        set_header_styles(ws['B6'])

        for i in range(len(top_products)):
            product_name = top_products[i]['product_varient__product__product_name']
            quantity = top_products[i]['total_quantity']

            ws[f'A{i+7}'] = product_name
            ws[f'B{i+7}'] = quantity

        today_order_products=OrderItems.objects.filter(order__in=todays_order)

        ws['A15']='Product'
        ws['B15']='Varient'
        ws['c15']='Price'
        ws['D15']='Order Id'

        set_header_styles(ws['A15'])
        set_header_styles(ws['B15'])
        set_header_styles(ws['C15'])
        set_header_styles(ws['D15'])


        for i, order_product in enumerate(today_order_products):
            ws[f'A{i+16}'] = order_product.product_varient.product.product_name
            ws[f'B{i+16}'] = str(order_product.product_varient.color) 
            ws[f'C{i+16}'] = order_product.price
            ws[f'D{i+16}'] = order_product.order.id


    elif time_period =='week':
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        week_order=Order.objects.filter(order_date__range=[week_ago, today])
        week_order_count=week_order.count()
        week_revenue = Order.objects.filter(order_date__range=[week_ago, today]).aggregate(week_revenue=Sum('total_amount'))['week_revenue']
        weekly_order_items=OrderItems.objects.filter(order__in=week_order).count()

        week_top_products = OrderItems.objects.filter(order__in=week_order).values(
            'product_varient__product__product_name',
        ).annotate(
            total_quantity=Sum('quantity')
        ).order_by('-total_quantity')[:5]

        ws['A1'] = 'Total Order'
        ws['A2'] = 'Total revenue'
        ws['A3'] = 'Total order items'
        ws['B1'] = week_order_count
        ws['B2'] = week_revenue
        ws['B3'] = weekly_order_items

        for cell in ws['A1:B3']:
            set_header_styles(cell[0])

        ws['A6']='Product'
        ws['B6']='Quantity'

        set_header_styles(ws['A6'])
        set_header_styles(ws['B6'])

        for i in range(len(week_top_products)):
            product_name = week_top_products[i]['product_varient__product__product_name']
            quantity = week_top_products[i]['total_quantity']

            ws[f'A{i+7}'] = product_name
            ws[f'B{i+7}'] = quantity
        
        weekly_order_products=OrderItems.objects.filter(order__in=week_order)

        ws['A15']='Product'
        ws['B15']='Varient'
        ws['c15']='Price'
        ws['D15']='Order Id'

        set_header_styles(ws['A15'])
        set_header_styles(ws['B15'])
        set_header_styles(ws['C15'])
        set_header_styles(ws['D15'])


        for i, order_product in enumerate(weekly_order_products):
            ws[f'A{i+16}'] = order_product.product_varient.product.product_name
            ws[f'B{i+16}'] = str(order_product.product_varient.color) 
            ws[f'C{i+16}'] = order_product.price
            ws[f'D{i+16}'] = order_product.order.id

        
    
    elif time_period == 'month':
        today = timezone.now().date()
        month_start = today.replace(day=1)
        month_order=Order.objects.filter(order_date__range=[month_start, today])
        month_order_count=month_order.count()
        month_revenue = Order.objects.filter(order_date__range=[month_start, today]).aggregate(month_revenue=Sum('total_amount'))['month_revenue']
        monthly_order_items=OrderItems.objects.filter(order__in=month_order).count()

        month_top_products = OrderItems.objects.filter(order__in=month_order).values(
            'product_varient__product__product_name','quantity'
        ).annotate(
            total_quantity=Sum('quantity')
        ).order_by('-total_quantity')[:5]

        

        ws['A1'] = 'Total Order'
        ws['A2'] = 'Total revenue'
        ws['A3'] = 'Total order items'
        ws['B1'] = month_order_count
        ws['B2'] = month_revenue
        ws['B3'] = monthly_order_items

        for cell in ws['A1:B3']:
            set_header_styles(cell[0])

        ws['A6']='Product'
        ws['B6']='Quantity'

        set_header_styles(ws['A6'])
        set_header_styles(ws['B6'])

        for i in range(len(month_top_products)):
            product_name = month_top_products[i]['product_varient__product__product_name']
            quantity = month_top_products[i]['total_quantity']

            ws[f'A{i+7}'] = product_name
            ws[f'B{i+7}'] = quantity

        monthly_order_products=OrderItems.objects.filter(order__in=month_order)

        ws['A15']='Product'
        ws['B15']='Varient'
        ws['c15']='Price'
        ws['D15']='Order Id'

        set_header_styles(ws['A15'])
        set_header_styles(ws['B15'])
        set_header_styles(ws['C15'])
        set_header_styles(ws['D15'])


        for i, order_product in enumerate(monthly_order_products):
            ws[f'A{i+16}'] = order_product.product_varient.product.product_name
            ws[f'B{i+16}'] = str(order_product.product_varient.color) 
            ws[f'C{i+16}'] = order_product.price
            ws[f'D{i+16}'] = order_product.order.id



        

    elif time_period == 'year':
        today = timezone.now().date()
        month_start = today.replace(day=1)
        year_start = today.replace(month=1, day=1)
        year_order = Order.objects.filter(order_date__range=[year_start, today])
        year_order_count = year_order.count()
        year_revenue = year_order.aggregate(year_revenue=Sum('total_amount'))['year_revenue']
        yearly_order_items = OrderItems.objects.filter(order__in=year_order).count()
        year_top_products = OrderItems.objects.filter(order__in=year_order).values(
            'product_varient__product__product_name', 'quantity'
        ).annotate(
            total_quantity=Sum('quantity')
        ).order_by('-total_quantity')[:5]

        ws['A1'] = 'Total Order'
        ws['A2'] = 'Total revenue'
        ws['A3'] = 'Total order items'
        ws['B1'] = year_order_count
        ws['B2'] = year_revenue
        ws['B3'] = yearly_order_items

        for cell in ws['A1:B3']:
            set_header_styles(cell[0])

        ws['A6']='Product'
        ws['B6']='Quantity'

        set_header_styles(ws['A6'])
        set_header_styles(ws['B6'])

        for i in range(len(year_top_products)):
            product_name = year_top_products[i]['product_varient__product__product_name']
            quantity = year_top_products[i]['total_quantity']

            ws[f'A{i+7}'] = product_name
            ws[f'B{i+7}'] = quantity

        yearly_order_products=OrderItems.objects.filter(order__in=year_order)
        
        ws['A15']='Product'
        ws['B15']='Varient'
        ws['c15']='Price'
        ws['D15']='Order Id'

        set_header_styles(ws['A15'])
        set_header_styles(ws['B15'])
        set_header_styles(ws['C15'])
        set_header_styles(ws['D15'])


        for i, order_product in enumerate(yearly_order_products):
            ws[f'A{i+16}'] = order_product.product_varient.product.product_name
            ws[f'B{i+16}'] = str(order_product.product_varient.color) 
            ws[f'C{i+16}'] = order_product.price
            ws[f'D{i+16}'] = order_product.order.id
    
    elif from_date and to_date:
        orders = Order.objects.filter(order_date__range=[from_date, to_date])
        order_count=orders.count()
        revenue = Order.objects.filter(order_date__range=[from_date, to_date]).aggregate(week_revenue=Sum('total_amount'))['week_revenue']

        order_items = OrderItems.objects.filter(order__in=orders).count()
        top_products = list(
            OrderItems.objects.filter(order__in=orders).values(
                'product_varient__product__product_name', 'quantity'
            ).annotate(
                total_quantity=Sum('quantity')
            ).order_by('-total_quantity')[:5]
        )
        ws['A1'] = 'Total Order'
        ws['A2'] = 'Total revenue'
        ws['A3'] = 'Total order items'
        ws['B1'] = order_count
        ws['B2'] = revenue
        ws['B3'] = order_items

        for cell in ws['A1:B3']:
            set_header_styles(cell[0])

        ws['A6']='Product'
        ws['B6']='Quantity'

        for cell in ws['A6:B6']:
            set_header_styles(cell)

        for i in range(len(top_products)):
            product_name = top_products[i]['product_varient__product__product_name']
            quantity = top_products[i]['total_quantity']

            ws[f'A{i+7}'] = product_name
            ws[f'B{i+7}'] = quantity
        
        order_products=OrderItems.objects.filter(order__in=orders)
        
        ws['A15']='Product'
        ws['B15']='Varient'
        ws['c15']='Price'
        ws['D15']='Order Id'

        set_header_styles(ws['A15'])
        set_header_styles(ws['B15'])
        set_header_styles(ws['C15'])
        set_header_styles(ws['D15'])


        for i, order_product in enumerate(order_products):
            ws[f'A{i+16}'] = order_product.product_varient.product.product_name
            ws[f'B{i+16}'] = str(order_product.product_varient.color) 
            ws[f'C{i+16}'] = order_product.price
            ws[f'D{i+16}'] = order_product.order.id
    


  

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=report.xlsx'
    wb.save(response)

    return response


